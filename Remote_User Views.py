from django.db.models import Count
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
import datetime
import openpyxl
import re
import string
from sklearn.feature_extraction.text import CountVectorizer

import pandas as pd

from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics import accuracy_score, confusion_matrix, classification_report
from sklearn.metrics import accuracy_score
from sklearn.metrics import f1_score
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import VotingClassifier


# Create your views here.
from Remote_User.models import ClientRegister_Model,Tweet_Message_details,detection_ratio,detection_accuracy,Emotion_prediction

def login(request):


    if request.method == "POST" and 'submit1' in request.POST:

        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            enter = ClientRegister_Model.objects.get(username=username,password=password)
            request.session["userid"] = enter.id

            return redirect('ViewYourProfile')
        except:
            pass

    return render(request,'RUser/login.html')

def Add_DataSet_Details(request):
    if "GET" == request.method:
        return render(request, 'RUser/Add_DataSet_Details.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)
        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)
        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)
        # reading a cell
        print(worksheet["A1"].value)
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)
            Tweet_Message_details.objects.all().delete()

    for r in range(1, active_sheet.max_row+1):
        Tweet_Message_details.objects.create(
        Tweet_Id= active_sheet.cell(r, 1).value,
        Tweet_Label= active_sheet.cell(r, 2).value,
        Tweet_Message= active_sheet.cell(r, 3).value,
        )

    return render(request, 'RUser/Add_DataSet_Details.html', {"excel_data": excel_data})


def Register1(request):

    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        ClientRegister_Model.objects.create(username=username, email=email, password=password, phoneno=phoneno,
                                            country=country, state=state, city=city)

        return render(request, 'RUser/Register1.html')
    else:
        return render(request,'RUser/Register1.html')

def ViewYourProfile(request):
    userid = request.session['userid']
    obj = ClientRegister_Model.objects.get(id= userid)
    return render(request,'RUser/ViewYourProfile.html',{'object':obj})


def Search_DataSets(request):
    if request.method == "POST":
        kword = request.POST.get('keyword')
        if request.method == "POST":
            kword = request.POST.get('keyword')
            User_ID= request.POST.get('uid')
            print(kword)
            data = pd.read_csv("Datasets.csv")

            def clean_text(text):
                '''Make text lowercase, remove text in square brackets,remove links,remove punctuation
                and remove words containing numbers.'''
                text = text.lower()
                text = re.sub('\[.*?\]', '', text)
                text = re.sub('https?://\S+|www\.\S+', '', text)
                text = re.sub('<.*?>+', '', text)
                text = re.sub('[%s]' % re.escape(string.punctuation), '', text)
                text = re.sub('\n', '', text)
                text = re.sub('\w*\d\w*', '', text)
                text = re.sub('@', '', text)
                text = re.sub('!', '', text)
                text = re.sub('#', '', text)
                return text

            data['text'] = data['tweet'].apply(lambda x: clean_text(x))

            def remove_emoji(text):
                emoji_pattern = re.compile("["
                                           u"\U0001F600-\U0001F64F"  # emoticons
                                           u"\U0001F300-\U0001F5FF"  # symbols & pictographs
                                           u"\U0001F680-\U0001F6FF"  # transport & map symbols
                                           u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
                                           u"\U00002702-\U000027B0"
                                           u"\U000024C2-\U0001F251"
                                           "]+", flags=re.UNICODE)
                return emoji_pattern.sub(r'', text)

            data['text'] = data['tweet'].apply(lambda x: remove_emoji(x))
            data['text'].apply(lambda x: len(str(x).split())).max()

            # Creating a mapping for Review Analysis

            x = data['tweet']
            y = data['label']
            cv = CountVectorizer()
            x = cv.fit_transform(x)

            models = []
            from sklearn.model_selection import train_test_split
            X_train, X_test, y_train, y_test = train_test_split(x, y, test_size=0.20)
            X_train.shape, X_test.shape, y_train.shape

            print("Naive Bayes")

            from sklearn.naive_bayes import MultinomialNB

            NB = MultinomialNB()
            NB.fit(X_train, y_train)
            predict_nb = NB.predict(X_test)
            naivebayes = accuracy_score(y_test, predict_nb) * 100
            print(naivebayes)
            print(confusion_matrix(y_test, predict_nb))
            print(classification_report(y_test, predict_nb))
            models.append(('naive_bayes', NB))

            # SVM Model
            print("SVM")
            from sklearn import svm

            lin_clf = svm.LinearSVC()
            lin_clf.fit(X_train, y_train)
            predict_svm = lin_clf.predict(X_test)
            svm_acc = accuracy_score(y_test, predict_svm) * 100
            print(svm_acc)
            print("CLASSIFICATION REPORT")
            print(classification_report(y_test, predict_svm))
            print("CONFUSION MATRIX")
            print(confusion_matrix(y_test, predict_svm))
            models.append(('svm', lin_clf))

            print("Logistic Regression")

            from sklearn.linear_model import LogisticRegression

            reg = LogisticRegression(random_state=0, solver='lbfgs').fit(X_train, y_train)
            y_pred = reg.predict(X_test)
            print("ACCURACY")
            print(accuracy_score(y_test, y_pred) * 100)
            print("CLASSIFICATION REPORT")
            print(classification_report(y_test, y_pred))
            print("CONFUSION MATRIX")
            print(confusion_matrix(y_test, y_pred))
            models.append(('logistic', reg))

            print("Decision Tree Classifier")
            dtc = DecisionTreeClassifier()
            dtc.fit(X_train, y_train)
            dtcpredict = dtc.predict(X_test)
            print("ACCURACY")
            print(accuracy_score(y_test, dtcpredict) * 100)
            print("CLASSIFICATION REPORT")
            print(classification_report(y_test, dtcpredict))
            print("CONFUSION MATRIX")
            print(confusion_matrix(y_test, dtcpredict))
            models.append(('DecisionTreeClassifier', dtc))

            print("SGD Classifier")
            from sklearn.linear_model import SGDClassifier
            sgd_clf = SGDClassifier(loss='hinge', penalty='l2', random_state=0)
            sgd_clf.fit(X_train, y_train)
            sgdpredict = sgd_clf.predict(X_test)
            print("ACCURACY")
            print(accuracy_score(y_test, sgdpredict) * 100)
            print("CLASSIFICATION REPORT")
            print(classification_report(y_test, sgdpredict))
            print("CONFUSION MATRIX")
            print(confusion_matrix(y_test, sgdpredict))
            models.append(('SGDClassifier', sgd_clf))

            classifier = VotingClassifier(models)
            classifier.fit(X_train, y_train)
            y_pred = classifier.predict(X_test)

            Reviews_data = [kword]
            vector1 = cv.transform(Reviews_data).toarray()
            predict_text = classifier.predict(vector1)

            pred = str(predict_text).replace("[", "")
            pred1 = pred.replace("]", "")

            prediction = int(pred1)

            if prediction == 0:
                val = 'Happy'
            elif prediction == 1:
                val = 'Un Happy'

            print(val);

            Emotion_prediction.objects.create(Tweet_Id=User_ID, Tweet_Message=kword,Emotion_Prediction=val)

        return render(request, 'RUser/Search_DataSets.html',{'objs': val})
    return render(request, 'RUser/Search_DataSets.html')



